# -*- mode:python; coding:utf-8 -*-
# Copyright (c) 2023 IBM Corp. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""OSCAL transformation tasks."""

import configparser
import datetime
import logging
import pathlib
import traceback
import uuid
from collections import OrderedDict
from typing import Iterator, List, Optional

from openpyxl import load_workbook

from trestle.oscal import OSCAL_VERSION
from trestle.oscal.catalog import Catalog
from trestle.oscal.catalog import Control
from trestle.oscal.catalog import Group
from trestle.oscal.common import Link
from trestle.oscal.common import Metadata
from trestle.oscal.common import Part
from trestle.oscal.common import Property
from trestle.tasks.base_task import TaskBase
from trestle.tasks.base_task import TaskOutcome

logger = logging.getLogger(__name__)

timestamp = datetime.datetime.utcnow().replace(microsecond=0).replace(tzinfo=datetime.timezone.utc).isoformat()


class XlsxHelper:
    """Xlsx Helper common functions and assistance navigating spread sheet."""

    def __init__(self, file: str) -> None:
        """Initialize."""
        self._spread_sheet = file
        self._wb = load_workbook(self._spread_sheet)
        self._sheet_name = 'Combined Profiles'
        if self._sheet_name not in self._wb.sheetnames:
            raise RuntimeError(f'{file} missing {self._sheet_name} sheet')
        self._work_sheet = self._wb[self._sheet_name]
        self._mapper()

    def _mapper(self) -> None:
        """Map columns heading names to column numbers."""
        self._col_name_to_number = {}
        cols = self._work_sheet.max_column
        row = 1
        for col in range(row, cols):
            cell = self._work_sheet.cell(row, col)
            if cell.value:
                name = cell.value
                self._col_name_to_number[name] = col

    def row_generator(self) -> Iterator[int]:
        """Generate rows until max reached."""
        row = 2
        while row <= self._work_sheet.max_row:
            yield row
            row += 1

    def get(self, row: int, name: str) -> str:
        """Get cell value for given row and column name."""
        col = self._col_name_to_number[name]
        cell = self._work_sheet.cell(row, col)
        return cell.value


class CatalogHelper:
    """OSCAL Catalog Helper."""

    def __init__(self, title: str, version: str) -> None:
        """Initialize."""
        # metadata
        self._metadata = Metadata(title=title, last_modified=timestamp, oscal_version=OSCAL_VERSION, version=version)
        self._root_group = OrderedDict()
        self._all_groups = OrderedDict()

    def add_group(self, section: str, title: str, props: List[Property], parts: List[Part]) -> None:
        """Add group."""
        numdots = section.count('.')
        if numdots == 0:
            group = Group(title=f'{title}', id=f'CIS-{section}')
            if len(props):
                group.props = props
            if len(parts):
                group.parts = parts
            self._root_group[section] = group
            self._all_groups[section] = group
        elif numdots == 1:
            key = section.split('.')[0]
            parent = self._root_group[key]
            if parent.groups is None:
                parent.groups = []
            group = Group(title=f'{title}', id=f'CIS-{section}')
            if len(props):
                group.props = props
            if len(parts):
                group.parts = parts
            parent.groups.append(group)
            self._all_groups[section] = group
        else:
            raise RuntimeError(f'unexpected section {section}')

    def add_control(
        self,
        section: str,
        recommendation: str,
        title: str,
        props: List[Property],
        parts: List[Part],
        links: List[Link]
    ) -> None:
        """Add control."""
        group = self._all_groups[section]
        if group.controls is None:
            group.controls = []
        id_ = f'CIS-{recommendation}'
        title = f'{title}'
        control = Control(id=id_, title=title)
        if len(props):
            control.props = props
        if len(parts):
            control.parts = parts
        if len(links):
            control.links = links
        group.controls.append(control)

    def get_catalog(self) -> Catalog:
        """Get catalog."""
        catalog = Catalog(uuid=str(uuid.uuid4()), metadata=self._metadata, groups=list(self._root_group.values()))
        return catalog


class CisXlsxToOscalCatalog(TaskBase):
    """
    Task to transform CIS .xlsx to OSCAL catalog.

    Attributes:
        name: Name of the task.
    """

    name = 'cis-xlsx-to-oscal-catalog'

    def __init__(self, config_object: Optional[configparser.SectionProxy]) -> None:
        """
        Initialize trestle task ocp4-cis-profile-to-oscal-catalog.

        Args:
            config_object: Config section associated with the task.
        """
        super().__init__(config_object)

    def print_info(self) -> None:
        """Print the help string."""
        logger.info(f'Help information for {self.name} task.')
        logger.info('')
        logger.info('Purpose: Create catalog from standard (e.g. CIS benchmark).')
        logger.info('')
        logger.info('Configuration flags sit under [task.cis-xlsx-to-oscal-catalog]:')
        text1 = '  input-file             = '
        text2 = '(required) path to read the compliance-as-code .xlsx spread sheet file.'
        logger.info(text1 + text2)
        text1 = '  output-dir             = '
        text2 = '(required) location to write the generated catalog.json file.'
        logger.info(text1 + text2)
        text1 = '  title                  = '
        text2 = '(required) title of the CIS catalog.'
        logger.info(text1 + text2)
        text1 = '  version                = '
        text2 = '(required) version :q!of the CIS catalog.'
        logger.info(text1 + text2)
        text1 = '  output-overwrite       = '
        text2 = '(optional) true [default] or false; replace existing output when true.'
        logger.info(text1 + text2)

    def simulate(self) -> TaskOutcome:
        """Provide a simulated outcome."""
        return TaskOutcome('simulated-success')

    def execute(self) -> TaskOutcome:
        """Provide an actual outcome."""
        try:
            return self._execute()
        except Exception:
            logger.info(traceback.format_exc())
            return TaskOutcome('failure')

    def _get_normalized_name(self, key: str) -> str:
        """Get normalized name."""
        name = key
        name = name.replace(' ', '_')
        name = name.replace('&', 'a')
        return name

    def _add_property(self, xlsx_helper: XlsxHelper, props: List[Property], row: int, key: str) -> None:
        """Add property."""
        name = self._get_normalized_name(key)
        value = xlsx_helper.get(row, key)
        if value:
            props.append(Property(name=name, value=value))

    def _add_property_boolean(self, xlsx_helper: XlsxHelper, props: List[Property], row: int, key: str) -> None:
        """Add property."""
        name = self._get_normalized_name(key)
        value = xlsx_helper.get(row, key)
        if value:
            props.append(Property(name=name, value='True'))
        else:
            props.append(Property(name=name, value='False'))

    def _add_part(self, xlsx_helper: XlsxHelper, parts: List[Part], id_: str, row: int, key: str) -> None:
        """Add part."""
        value = xlsx_helper.get(row, key)
        if value:
            name = self._get_normalized_name(key)
            parts.append(Part(id=id_, name=name, prose=value))

    def _add_links(self, xlsx_helper: XlsxHelper, links: List[Link], row: int, key: str) -> None:
        """Add links."""
        value = xlsx_helper.get(row, key)
        if value:
            value = value.replace(':http', ' http')
            urls = value.split()
            for url in urls:
                link = Link(href=url, rel='reference')
                links.append(link)

    def _execute(self) -> TaskOutcome:
        """Wrap the execute for exception handling."""
        if not self._config:
            logger.warning('config missing')
            return TaskOutcome('failure')
        try:
            ifile = self._config['input-file']
            odir = self._config['output-dir']
            title = self._config['title']
            version = self._config['version']
        except KeyError as e:
            logger.info(f'key {e.args[0]} missing')
            return TaskOutcome('failure')
        # verbosity
        _quiet = self._config.get('quiet', False)
        _verbose = not _quiet
        # output
        overwrite = self._config.getboolean('output-overwrite', True)
        opth = pathlib.Path(odir)
        # insure output dir exists
        opth.mkdir(exist_ok=True, parents=True)
        # calculate output file name & check writability
        oname = 'catalog.json'
        ofile = opth / oname
        if not overwrite and pathlib.Path(ofile).exists():
            logger.warning(f'output: {ofile} already exists')
            return TaskOutcome('failure')
        xlsx_helper = XlsxHelper(ifile)
        catalog_helper = CatalogHelper(title, version)
        # transform each row into OSCAL equivalent
        for row in xlsx_helper.row_generator():
            section = xlsx_helper.get(row, 'section #')
            recommendation = xlsx_helper.get(row, 'recommendation #')
            title = xlsx_helper.get(row, 'title')
            # init
            props = []
            parts = []
            links = []
            # props
            self._add_property(xlsx_helper, props, row, 'profile')
            self._add_property(xlsx_helper, props, row, 'status')
            self._add_property(xlsx_helper, props, row, 'assessment status')
            # parts
            self._add_part(xlsx_helper, parts, f'CIS-{section}_des', row, 'description')
            self._add_part(xlsx_helper, parts, f'CIS-{section}_rat', row, 'rationale statement')
            self._add_part(xlsx_helper, parts, f'CIS-{section}_imp', row, 'impact statement')
            self._add_part(xlsx_helper, parts, f'CIS-{section}_rem', row, 'remediation procedure')
            self._add_part(xlsx_helper, parts, f'CIS-{section}_aud', row, 'audit procedure')
            self._add_part(xlsx_helper, parts, f'CIS-{section}_inf', row, 'additional information')
            self._add_part(xlsx_helper, parts, f'CIS-{section}_ctl', row, 'CIS Controls')
            # group or control
            if recommendation is None:
                catalog_helper.add_group(section, title, props, parts)
            else:
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG1')
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG2')
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG3')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG1')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG2')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG3')
                self._add_property(xlsx_helper, props, row, 'MITRE ATT&CK Mappings')
                self._add_links(xlsx_helper, links, row, 'references')
                catalog_helper.add_control(section, recommendation, title, props, parts, links)
        catalog = catalog_helper.get_catalog()
        # write OSCAL ComponentDefinition to file
        if _verbose:
            logger.info(f'output: {ofile}')
        catalog.oscal_write(pathlib.Path(ofile))
        return TaskOutcome('success')